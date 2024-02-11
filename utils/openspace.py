from table import Table
from table import Seat
import random
import openpyxl

class Openspace:
    """
    Initialize an Openspace object with a specified number of tables.

    Paramenters:
        number_of_tables (int): The number of tables in the openspace. Default is 6.
    """
    def __init__(self, number_of_tables: int = 6, table_capacity: int = 4):
        self.tables = [Table(table_capacity) for i in range(number_of_tables)]

    
    def __str__(self):
        return f"Openspace has {len(self.tables)} tables."

    def add_table(self, how_many_add=1, table_capacity=4):
        self.tables.extend([Table(table_capacity) for i in range(how_many_add)])

    def empty_all_tables(self):
        for table in self.tables:
            for seat in table.seats:
                seat.remove_occupant()
    
    def organize(self, names: list[str]):
        """
        Organize the seating arrangement for a list of names.
        Paramenters:
        names (list[str]): The list of names to be seated.
        """
        for name in names:
            self.assign_seat(name)

    def has_free_table(self) -> bool:
        """
        Check if there is at least one free table in the openspace.
        Returns:
        bool: True if there is a free table, False otherwise.
        """
        for table in self.tables:
            if table.has_free_spot():
                return True
        return False
 
    def assign_seat(self, name: str):
        """
        Assigns a seat to a person with the given name.
        Paramenters:
        name (str): The name of the person to be seated.
        Raises:
        Exception: If all the tables are occupied.
        """
        if self.has_free_table():
            random.shuffle(self.tables)
            for table in self.tables:
                if table.has_free_spot() == True:
                    table.assign_seat(name)
                    break
        else:
            raise Exception("All the tables are occupied")

    def has_free_spot_for_group(self, group_size: int) -> bool:
        """
        Checks if there is a free table with enough capacity for a group of a given size.
        Paramenters:
        group_size (int): The size of the group.
        Returns:
         bool: True if there is a free table with enough capacity, False otherwise.
        """
        for table in self.tables:
            if table.left_capacity() >= group_size:
                return True
        return False


    def assign_seat_for_group(self, names_in_group: list[str]):
        """
        Assigns seats to a group of people with the given names.
        Args:
        names_in_group (list[str]): The list of names in the group.
        Raises:
        Exception: If all the tables are occupied.
        """
        group_size = len(names_in_group)
        if self.has_free_spot_for_group(group_size) == True:
            for table in self.tables:
                if table.left_capacity()>= group_size:
                    for name in names_in_group:
                        table.assign_seat(name)
                    break
        else:
            raise Exception("Not many seats left, sit separately")

    def has_free_spot_for_group_with_rivals(self, group_size, names_of_rivals):
        """
        Checks if rivals are not at the table.
        Paramenters:
        group_size (int): The size of the group.
        names_of_rivals (list[str]): The list of names of people whom someone else named at rivals.
        bool: True if there is a free table with enough capacity, False otherwise.
        """
        for table in self.tables:
            occupants = [seat.occupant for seat in table.seats if seat.free == False]
            rivals_not_at_table = True # pending evidence...
            for rival in names_of_rivals:
                if rival in occupants:
                    rivals_not_at_table = False
            if table.left_capacity() >= group_size and rivals_not_at_table:
                return True
        return False

    def assign_seat_for_group_with_rivals(self, names_in_group, names_of_rivals):
        """
        Assign seats to a group of people with the given names, and makes sure that they do not sit with their rivals.
        Args:
        names_in_group (list[str]): The list of names in the group
        names_of_rivals (list[str]): The list of names of people with whonm the first group did not want to sit
        Raises:
        Exception: If all the tables are occupied.
        """
        group_size = len(names_in_group)
        if self.has_free_spot_for_group_with_rivals(group_size, names_of_rivals) == True:
            for table in self.tables:
                occupants = [seat.occupant for seat in table.seats if seat.free == False]
                rivals_not_at_table = True # pending evidence...
                for rival in names_of_rivals:
                    if rival in occupants:
                        rivals_not_at_table = False
                if table.left_capacity() >= group_size and rivals_not_at_table:
                    for name in names_in_group:
                        table.assign_seat(name)
                    break
        else:
            raise Exception("not enough tables")



    def display (self):
        """
        Displays the seating arrangement in the openspace.

        Example output:
        a, b, c are seated at the same table
        ...
        x, y, z are seated at the same table
        """
        
        for table in self.tables:
            occupants = [seat.occupant for seat in table.seats if seat.free == False]
            if table.left_capacity() ==(table.capacity-1):
                print(f"{occupants[0]} is sitting alone.")
            if table.left_capacity() <(table.capacity-1):
                print (", ".join(occupants)+" are seated at the same table.")
    


    def store_occupants_in_excel(self, filename = "Openspace"):
        """
        Store the occupants of each table in an Excel document.

        Parameters:
        filename (str): The name of the Excel file to store the data.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        for i, table in enumerate(self.tables, start=1):
            occupants = [seat.occupant for seat in table.seats if not seat.free]
            occupants_str = ", ".join(occupants)
            sheet.cell(row=i, column=1, value=f"Table {i}")
            sheet.cell(row=i, column=2, value=occupants_str)

        workbook.save(filename+".xlsx")

